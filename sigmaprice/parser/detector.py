"""File format and structure detection"""
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional
import openpyxl
import chardet
from sigmaprice.core.exceptions import ParseError
from sigmaprice.core.logger import get_logger

logger = get_logger(__name__)


def detect_format(file_path: Path) -> str:
    """
    Detect file format from extension and content.

    Args:
        file_path: Path to file

    Returns:
        'xlsx', 'xls', or 'csv'

    Raises:
        ParseError: If format not supported
    """
    ext = file_path.suffix.lower()

    if ext in ['.xlsx', '.xls']:
        return ext[1:]  # 'xlsx' or 'xls'

    if ext == '.csv':
        return 'csv'

    raise ParseError(f"Unsupported file format: {ext}")


def detect_encoding(file_path: Path) -> str:
    """
    Detect file encoding for CSV.

    Args:
        file_path: Path to CSV file

    Returns:
        Detected encoding (utf-8, windows-1251, etc.)
    """
    try:
        with open(file_path, 'rb') as f:
            result = chardet.detect(f.read(10000))
            encoding = result['encoding'] or 'utf-8'
            logger.info(f"Detected encoding: {encoding}")
            return encoding
    except Exception as e:
        logger.warning(f"Failed to detect encoding, defaulting to utf-8: {e}")
        return 'utf-8'


def detect_delimiter(file_path: Path, encoding: str = 'utf-8') -> str:
    """
    Detect CSV delimiter.

    Args:
        file_path: Path to CSV file
        encoding: File encoding

    Returns:
        Delimiter character (, ; \t)
    """
    try:
        with open(file_path, 'r', encoding=encoding, newline='') as f:
            sample = f.read(4096)

        comma = sample.count(',')
        semicolon = sample.count(';')
        tab = sample.count('\t')

        if semicolon > comma and semicolon >= tab:
            return ';'
        if tab > comma and tab >= semicolon:
            return '\t'
        return ','

    except Exception as e:
        logger.warning(f"Failed to detect delimiter, using comma: {e}")
        return ','


def detect_header_row(sheet_data: List[List[Any]]) -> int:
    """
    Determine which row contains headers.

    Heuristics:
    - Headers often contain: Код, Наименование, Цена, Артикул, etc.
    - Headers are usually in first 5 rows
    - Header cells are strings, data cells often numbers
    - First row with text matching known header keywords

    Args:
        sheet_data: List of rows (each row is list of cells)

    Returns:
        Row index (0-based) or 0 if not found (default to first row)
    """
    if not sheet_data:
        return 0

    header_keywords = [
        'код', 'наименование', 'цена', 'артикул', 'ean', 'upc',
        'наличие', 'количество', 'гарантия', 'производитель',
        'vendor', 'product', 'price', 'code', 'article', 'name'
    ]

    for row_idx, row in enumerate(sheet_data[:5]):
        if not row:
            continue

        text_values = [str(cell).lower().strip() if cell else '' for cell in row]
        keyword_count = sum(1 for tv in text_values if any(kw in tv for kw in header_keywords))

        if keyword_count >= 2:
            logger.info(f"Header found at row {row_idx}")
            return row_idx

    logger.warning("Could not detect header row, using first row")
    return 0


def detect_column_mapping(headers: List[str]) -> Dict[str, str]:
    """
    Auto-detect column mapping from header names.

    Args:
        headers: List of column headers from price file

    Returns:
        Dict of our_field -> supplier_column
    """
    mapping = {}
    normalized = {h.lower().strip(): h for h in headers if h}

    field_patterns = {
        'code': ['код', 'sku', 'id товара', 'код товара', 'article id'],
        'name': ['наименование', 'название', 'товар', 'product', 'name', 'наименование товара'],
        'description': ['описание', 'desc', 'детали', 'детальное описание'],
        'price': ['цена', 'price', 'стоимость', 'сумма', 'розничная'],
        'availability': ['наличие', 'статус', 'availability', 'в наличии', 'остаток'],
        'quantity': ['количество', 'qty', 'остаток', 'кол-во', 'на складе'],
        'warranty_months': ['гарантия', 'warranty', 'гарантия (мес)'],
        'article': ['артикул', 'модель', 'part number', 'pn', 'арт'],
        'ean': ['ean', 'upc', 'штрих', 'barcode', 'штрих-код'],
        'manufacturer': ['производитель', 'brand', 'vendor', 'изготовитель', 'бренд'],
        'delivery_type': ['тип', 'поставка', 'delivery', 'тип поставки', 'тип товара'],
        'category': ['категория', 'группа', 'category', 'раздел'],
    }

    for field, patterns in field_patterns.items():
        for pattern in patterns:
            for norm_header, original_header in normalized.items():
                if pattern in norm_header:
                    mapping[field] = original_header
                    break
            if field in mapping:
                break

    logger.info(f"Auto-detected mapping: {mapping}")
    return mapping


def get_all_sheets(file_path: Path) -> List[Tuple[str, bool]]:
    """
    Get list of all sheets in Excel file.

    Args:
        file_path: Path to Excel file

    Returns:
        List of tuples (sheet_name, is_visible)
    """
    try:
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheets = []

        for name in workbook.sheetnames:
            sheet = workbook[name]
            is_visible = sheet.sheet_state == 'visible'
            sheets.append((name, is_visible))

        workbook.close()
        return sheets

    except Exception as e:
        logger.error(f"Failed to read sheets: {e}")
        raise ParseError(f"Failed to read Excel file: {e}")


def is_sheet_empty(sheet) -> bool:
    """Check if sheet has no data rows."""
    for row in sheet.iter_rows(max_row=10, values_only=True):
        if any(cell for cell in row):
            return False
    return True