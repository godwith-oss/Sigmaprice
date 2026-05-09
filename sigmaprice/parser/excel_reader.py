"""Excel file reader for price lists"""
from pathlib import Path
from typing import Dict, List, Optional
import openpyxl
from sigmaprice.core.exceptions import ParseError
from sigmaprice.core.logger import get_logger

logger = get_logger(__name__)


def read_excel(
    file_path: Path,
    column_map: Optional[Dict[str, str]] = None
) -> List[Dict[str, str]]:
    """
    Read Excel file (xlsx/xls) and return list of rows with mapped columns.

    Args:
        file_path: Path to Excel file
        column_map: Dict of our_field -> supplier_column
            If provided, data will be mapped to our field names.
            If not provided, returns raw column names.

    Returns:
        List of dictionaries with normalized keys

    Raises:
        ParseError: If file cannot be read
    """
    if not file_path.exists():
        raise ParseError(f"File not found: {file_path}")

    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        raise ParseError(f"Failed to open Excel file: {e}")

    all_rows = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        if sheet.sheet_state != 'visible':
            logger.info(f"Skipping hidden sheet: {sheet_name}")
            continue

        logger.info(f"Processing sheet: {sheet_name}")

        sheet_data = _read_sheet_data(sheet)
        if not sheet_data:
            logger.warning(f"Empty sheet: {sheet_name}")
            continue

        header_row_idx = _find_header_row(sheet_data)
        headers = sheet_data[header_row_idx]

        reverse_map = {}
        if column_map:
            reverse_map = {v: k for k, v in column_map.items()}

        for row_idx in range(header_row_idx + 1, len(sheet_data)):
            row = sheet_data[row_idx]
            if not row or not any(cell for cell in row):
                continue

            row_dict = {'name': '', 'price': 0}

            for col_idx, header in enumerate(headers):
                if col_idx >= len(row):
                    continue

                value = row[col_idx]

                key = header.strip() if header else f'col_{col_idx}'
                if not key:
                    continue

                if column_map:
                    mapped_key = reverse_map.get(key, key)
                else:
                    mapped_key = key

                if value is not None:
                    row_dict[mapped_key] = value

            # Check if row has any data (not just empty)
            has_data = any(v for k, v in row_dict.items() if k not in ['name', 'price'] or v)
            if has_data and (row_dict.get('name') or row_dict.get('price') or any(row_dict.get(h) for h in headers)):
                all_rows.append(row_dict)

    workbook.close()
    logger.info(f"Total rows read: {len(all_rows)}")
    return all_rows


def _read_sheet_data(sheet) -> List[List]:
    """Read all data from sheet as list of lists."""
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))
    return data


def _find_header_row(sheet_data: List[List]) -> int:
    """Find row index with headers."""
    if not sheet_data:
        return 0

    header_keywords = [
        'код', 'наименование', 'цена', 'артикул', 'ean', 'upc',
        'наличие', 'количество', 'гарантия', 'производитель',
        'product', 'price', 'code', 'article', 'name'
    ]

    for row_idx, row in enumerate(sheet_data[:5]):
        if not row:
            continue

        text_values = [str(cell).lower().strip() if cell else '' for cell in row]
        keyword_count = sum(1 for tv in text_values if any(kw in tv for kw in header_keywords))

        if keyword_count >= 2:
            return row_idx

    return 0


def get_sheet_names(file_path: Path) -> List[str]:
    """Get list of sheet names in Excel file."""
    workbook = openpyxl.load_workbook(file_path, read_only=True)
    return workbook.sheetnames