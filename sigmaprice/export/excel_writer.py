"""Excel writer for catalog export"""
from pathlib import Path
from typing import List, Dict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from sigmaprice.core.logger import get_logger

logger = get_logger(__name__)

STATIC_COLUMNS = [
    ("Категория", 25),
    ("Производитель", 20),
    ("Резерв 1", 12),
    ("Резерв 2", 12),
    ("Код", 10),
    ("Наименование", 50),
    ("Описание", 40),
    ("Наша цена (₽)", 15),
    ("РРЦ (₽)", 15),
    ("Гарантия (мес)", 12),
    ("Артикул", 15),
    ("EAN/UPC", 15),
]

HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
RESERVE_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
PRICE_ALIGNMENT = Alignment(horizontal="right")
PRICE_FORMAT = '#,##0.00'
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)


def write_excel(
    rows: List[Dict],
    supplier_headers: List[str],
    filepath: Path,
) -> Path:
    """
    Write export data to an Excel file with formatting.

    Args:
        rows: List of row dicts from builder
        supplier_headers: List of supplier column names
        filepath: Output file path

    Returns:
        Path to created file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Каталог"

    all_headers = [col[0] for col in STATIC_COLUMNS] + supplier_headers

    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for col_idx, (name, width) in enumerate(STATIC_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for col_idx in range(len(STATIC_COLUMNS) + 1, len(all_headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, (col_name, _) in enumerate(STATIC_COLUMNS, start=1):
            value = row_data.get(col_name.lower().replace(" ", "_").replace("(", "").replace(")", "").replace("₽", ""))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.font = Font(name="Arial", size=9)

            if col_name in ("Резерв 1", "Резерв 2"):
                cell.fill = RESERVE_FILL

            if col_name in ("Наша цена (₽)", "РРЦ (₽)"):
                cell.alignment = PRICE_ALIGNMENT
                if value is not None:
                    cell.number_format = PRICE_FORMAT

            if col_name == "Код":
                cell.number_format = '@'

            if col_name == "Гарантия (мес)":
                cell.alignment = Alignment(horizontal="center")

        for i, supplier_header in enumerate(supplier_headers):
            col_idx = len(STATIC_COLUMNS) + i + 1
            is_price = "(цена)" in supplier_header
            key_base = f"supplier_{i // 2 + 1}"

            if is_price:
                value = row_data.get(f"{key_base}_price")
            else:
                value = row_data.get(f"{key_base}_available")

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.font = Font(name="Arial", size=9)

            if is_price and value is not None:
                cell.alignment = PRICE_ALIGNMENT
                cell.number_format = PRICE_FORMAT

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(all_headers))}{len(rows) + 1}"

    ws.sheet_properties.pageSetUpPr = None

    wb.save(filepath)
    logger.info(f"Excel export saved: {filepath} ({len(rows)} rows)")
    return filepath
